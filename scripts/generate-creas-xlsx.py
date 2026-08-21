#!/usr/bin/env python3
"""Génère l'Excel de production de créas depuis un JSON de lignes.

Usage :
    python3 scripts/generate-creas-xlsx.py <batch.json> <sortie.xlsx>

Le JSON d'entrée est une liste d'objets ; chaque objet est une créa dont les
clés correspondent aux colonnes ci-dessous (clé = `key`). Les clés absentes
laissent la cellule vide. Les colonnes et leurs valeurs autorisées viennent de
la formation (docs/formation/REGLES.md §3-4, docs/creas/PLAYBOOK.md,
docs/creas/CODIFICATION.md) — chaque colonne cite sa source dans la ligne 2 de
l'onglet Légende.

L'Excel produit contient :
  - onglet « Créas »   : une ligne par créa, filtres auto, listes déroulantes
  - onglet « Batchs »  : une ligne par batch (composition T36 rappelée)
  - onglet « Légende » : chaque colonne, sa définition, sa source formation
"""
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Le schéma : (clé, titre, largeur, valeurs autorisées ou None, source) ──
ANGLES = ["VEN", "FEM", "VAC", "PRO", "CNF", "STY"]
FORMATS = ["VID", "UGC", "IMG", "CAR"]
HOOKS = ["QST", "CHC", "PRB", "CUR", "TMG", "DEM"]
CONSCIENCE = ["unaware", "problem", "solution", "product", "most"]
LANGUES = ["FR", "EN", "ES", "DE"]
PAGES = ["marque", "tierce"]
STATUTS = ["brief", "généré", "monté", "lancé", "jugé"]
VERDICTS = ["", "WIN", "POTENTIAL", "SIGNAL", "ZOMBIE", "OFF"]
RATIOS = ["4:5", "9:16", "1:1"]

COLS = [
    ("code", "Code ad", 26, None,
     "AD<seq>v<var> - <PROD>-<ANGLE>-<FMT>-<HOOK>-<LANG> (CODIFICATION §3, T36 [03:59])"),
    ("batch", "Batch", 10, None, "B<AAAA-SS> (CODIFICATION §4)"),
    ("produit", "Produit", 10, ["POLO", "LANCASTER"], "registre produits du dashboard"),
    ("angle", "Angle", 8, ANGLES, "codes CODIFICATION §3 ; angles = review mining + winners (T08, PLAYBOOK angles)"),
    ("conscience", "Conscience", 11, CONSCIENCE, "5 stages d'awareness (T02-T05, PLAYBOOK angles)"),
    ("format", "Format", 8, FORMATS, "diversity map vidéo/statique (PLAYBOOK formats)"),
    ("style", "Style (diversity map)", 20, None, "style précis : talking head, unboxing, tableau noir… (PLAYBOOK formats)"),
    ("ratio", "Ratio", 7, RATIOS, "aspect ratios à tester (PLAYBOOK formats)"),
    ("hook_type", "Type hook", 9, HOOKS, "typologies hooks (CI-24, T17, PLAYBOOK hooks)"),
    ("hook_texte", "Hook (texte/visuel exact)", 46, None, "les 3 premières secondes, rédigées (PLAYBOOK hooks)"),
    ("script", "Script complet / brief visuel", 64, None, "blocs marketing dans l'ordre du framework choisi (T09-T11, PLAYBOOK scripts)"),
    ("framework", "Framework script", 16, None, "framework utilisé (winner 500K+, discredit, mini-VSL… — PLAYBOOK scripts)"),
    ("montage", "Consignes montage", 40, None, "rythme, coupes, sous-titres, b-roll, musique (PLAYBOOK formats)"),
    ("miniature", "Miniature (choisie à la main)", 30, None, "communique l'angle ; jamais l'auto (T36 [07:44])"),
    ("adcopy", "Ad copy (code + texte)", 46, None, "AC-<ANGLE>-<n> ; une adcopy par angle, Meta la lit (T36 [03:15], [10:22])"),
    ("titre", "Titre", 30, None, "TI-<ANGLE>-<n> ; 2-3 titres par batch (T36 [02:48])"),
    ("description", "Description", 26, None, "1 description par batch (T36 [02:48])"),
    ("page", "Page", 8, PAGES, "50 % marque / 50 % tierce (T36 [05:48])"),
    ("decline_de", "Décliné de (winner)", 18, None, "ad source si déclinaison ; règle 3-sur-5 : ≥ 3 éléments changés (T42 [08:48])"),
    ("elements_changes", "3-sur-5 : éléments changés", 24, None, "hook / visuel / texte / audio / format-message (T42 [08:48])"),
    ("statut", "Statut", 9, STATUTS, "cycle de production (Extension)"),
    ("date_lancement", "Lancé le", 11, None, "mardi→vendredi, jamais lundi ; live 00h-07h (T36 [00:20], [03:59])"),
    ("adset", "Adset", 26, None, "Creative Testing <mois> <sem>, blindé →50 (T36 [03:59], T42 [01:04])"),
    ("spend", "Spend €", 9, None, "résultat (rempli au jugement)"),
    ("roas", "ROAS", 7, None, "résultat"),
    ("marge", "Marge %", 8, None, "cm − 1/ROAS (backend, T38)"),
    ("ventes", "Ventes", 7, None, "résultat"),
    ("verdict", "Verdict", 10, VERDICTS,
     "WIN ≥6 ventes & ≥10 % · POTENTIAL entre BE et 10 % · SIGNAL <6 ventes & ≥15 % · ZOMBIE ≥6 sous BE (T37)"),
    ("notes", "Notes", 30, None, "libre"),
]

HEAD_FILL = PatternFill("solid", fgColor="1a1a2e")
HEAD_FONT = Font(color="7CFC9A", bold=True, size=10)
RESULT_FILL = PatternFill("solid", fgColor="f4ecd8")


def main(src: str, dst: str) -> None:
    rows = json.load(open(src, encoding="utf-8"))
    wb = Workbook()

    ws = wb.active
    ws.title = "Créas"
    for c, (_, title, width, _, _) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.fill, cell.font = HEAD_FILL, HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = "C2"
    ws.row_dimensions[1].height = 28

    for r, row in enumerate(rows, 2):
        for c, (key, _, _, _, _) in enumerate(COLS, 1):
            cell = ws.cell(row=r, column=c, value=row.get(key, ""))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if key in ("spend", "roas", "marge", "ventes", "verdict"):
                cell.fill = RESULT_FILL

    last = max(len(rows) + 40, 60)
    for c, (_, _, _, choices, _) in enumerate(COLS, 1):
        if choices:
            dv = DataValidation(
                type="list", formula1='"' + ",".join(choices) + '"', allow_blank=True
            )
            ws.add_data_validation(dv)
            col = get_column_letter(c)
            dv.add(f"{col}2:{col}{last}")
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{last}"

    wsb = wb.create_sheet("Batchs")
    wsb.append(["Batch", "Adset", "Nb ads", "Adcopies", "Titres", "Descriptions",
                "Pages marque/tierce", "Lancé le", "Conforme T36 ?"])
    for c in range(1, 10):
        wsb.cell(row=1, column=c).fill = HEAD_FILL
        wsb.cell(row=1, column=c).font = HEAD_FONT
        wsb.column_dimensions[get_column_letter(c)].width = 18
    wsb.append(["", "", "3-6", "2-3", "2-3", "1", "50/50", "mar→ven", ""])
    wsb.cell(row=2, column=1, value="(règle T36)")
    batches = {}
    for row in rows:
        batches.setdefault(row.get("batch", "?"), []).append(row)
    for b, items in sorted(batches.items()):
        pages = sum(1 for i in items if i.get("page") == "marque")
        wsb.append([b, items[0].get("adset", ""), len(items), "", "", "",
                    f"{pages}/{len(items) - pages}",
                    items[0].get("date_lancement", ""), ""])

    wsl = wb.create_sheet("Légende")
    wsl.append(["Colonne", "Définition / valeurs", "Source formation"])
    for c in range(1, 4):
        wsl.cell(row=1, column=c).fill = HEAD_FILL
        wsl.cell(row=1, column=c).font = HEAD_FONT
    wsl.column_dimensions["A"].width = 26
    wsl.column_dimensions["B"].width = 46
    wsl.column_dimensions["C"].width = 80
    for key, title, _, choices, source in COLS:
        wsl.append([title, " · ".join(choices) if choices else "", source])
        wsl.cell(row=wsl.max_row, column=3).alignment = Alignment(wrap_text=True)

    wb.save(dst)
    print(f"{dst} : {len(rows)} créas, {len(batches)} batch(s), {len(COLS)} colonnes")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    main(sys.argv[1], sys.argv[2])
